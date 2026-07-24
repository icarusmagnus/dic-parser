"""Parser EPT (European PRIIPs Template) — le format machine du DIC.

L'EPT est un fichier Excel/CSV standardisé (FinDatEx) qui contient EXACTEMENT
les données du DIC : ISIN, SRI, scénarios, et la ventilation complète des coûts.
Le parser ci-dessous, c'est le chemin FIABLE : pas de regex sur du PDF, on lit
des colonnes. Une ligne EPT = un partsclass (une part de fonds), identifiée par ISIN.

IMPORTANT — versions de gabarit :
Les codes de champ EPT ("01010", "05010"…) et surtout leurs libellés changent
entre EPT v1.1 / v2.0 / v3.0. Plutôt que de coder en dur des codes numériques
(fragiles), on matche les colonnes par MOTS-CLÉS dans l'en-tête, ce qui marche
sur les gabarits FR comme EN. Les indices `FIELD_CODE_HINTS` sont fournis en
secours : VÉRIFIE-les contre un EPT réel de ta source avant de t'y fier.
"""
from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Optional

from .models import KIDData, CostBreakdown, PerformanceScenario
from .utils import parse_number, parse_pct, valid_isin

try:
    from openpyxl import load_workbook
    _HAS_XLSX = True
except ImportError:  # openpyxl optionnel si tu ne lis que du CSV
    _HAS_XLSX = False


# Matching des colonnes par mots-clés (insensible casse/accents/espaces).
# Ordre = priorité : le premier motif qui matche une colonne gagne.
COLUMN_MATCHERS: dict[str, list[str]] = {
    "isin":              [r"\bisin\b", r"product.?identifier", r"^01010"],
    "product_name":      [r"priip.*name", r"product.*name", r"fund.*name", r"portfolio.*name"],
    "manufacturer":      [r"manufacturer", r"initiateur", r"management.*company"],
    "currency":          [r"\bcurrency\b", r"devise", r"share.?class.?currency"],
    "sri":               [r"\bsri\b", r"summary.?risk", r"risk.?indicator"],
    "rhp_years":         [r"recommended.?holding", r"\brhp\b", r"holding.?period.*year"],
    "entry_costs":       [r"entry.?cost", r"one.?off.*entry", r"coût.*entrée"],
    "exit_costs":        [r"exit.?cost", r"one.?off.*exit", r"coût.*sortie"],
    "ongoing_costs":     [r"ongoing.?cost", r"management.?fee", r"recurring.*management", r"coût.*récurrent"],
    "transaction_costs": [r"transaction.?cost", r"coût.*transaction"],
    "performance_fees":  [r"performance.?fee", r"commission.*résultat"],
    "carried_interest":  [r"carried.?interest", r"intéressement"],
    "riy_1y_pct":        [r"riy.*1.?y", r"reduction.?in.?yield.*1", r"total.?cost.*1.?y"],
    "riy_rhp_pct":       [r"riy.*rhp", r"reduction.?in.?yield.*rhp", r"total.?cost.*rhp"],
    # scénarios : rendement annuel moyen à la RHP
    "scen_stress_rhp":   [r"stress.*rhp.*(return|yield|annual)", r"stress.*annual.*return"],
    "scen_unfav_rhp":    [r"unfavou?rable.*rhp.*(return|yield|annual)"],
    "scen_moderate_rhp": [r"moderate.*rhp.*(return|yield|annual)"],
    "scen_favor_rhp":    [r"favou?rable.*rhp.*(return|yield|annual)"],
}

# Secours : mapping code EPT -> champ. À VÉRIFIER contre ton gabarit réel.
FIELD_CODE_HINTS: dict[str, str] = {
    "01010": "isin",
    "01030": "product_name",
    # ... complète avec les codes exacts de TON EPT (v2.0/v3.0) après vérification
}


def _norm(s: str) -> str:
    s = (s or "").lower()
    for a, b in (("é", "e"), ("è", "e"), ("ê", "e"), ("à", "a"), ("î", "i"), ("ï", "i")):
        s = s.replace(a, b)
    return re.sub(r"\s+", " ", s).strip()


def _build_column_index(header: list[str]) -> dict[str, int]:
    """Associe chaque champ logique à l'indice de colonne du fichier."""
    idx: dict[str, int] = {}
    normed = [_norm(h) for h in header]
    for field_name, patterns in COLUMN_MATCHERS.items():
        for col_i, h in enumerate(normed):
            if any(re.search(p, h) for p in patterns):
                idx.setdefault(field_name, col_i)
                break
    # secours par code EPT si un champ manque
    for code, field_name in FIELD_CODE_HINTS.items():
        if field_name not in idx:
            for col_i, h in enumerate(normed):
                if h.startswith(code):
                    idx[field_name] = col_i
                    break
    return idx


def _row_to_kid(row: list, cidx: dict[str, int], source_ref: str) -> Optional[KIDData]:
    def cell(name):
        i = cidx.get(name)
        return row[i] if i is not None and i < len(row) else None

    isin = str(cell("isin") or "").strip()
    if not valid_isin(isin):
        return None

    costs = CostBreakdown(
        entry_costs=parse_pct(cell("entry_costs")),
        exit_costs=parse_pct(cell("exit_costs")),
        ongoing_costs=parse_pct(cell("ongoing_costs")),
        transaction_costs=parse_pct(cell("transaction_costs")),
        performance_fees=parse_pct(cell("performance_fees")),
        carried_interest=parse_pct(cell("carried_interest")),
        riy_1y_pct=parse_pct(cell("riy_1y_pct")),
        riy_rhp_pct=parse_pct(cell("riy_rhp_pct")),
    )

    scenarios = []
    for key, label in [("scen_stress_rhp", "Tendu"), ("scen_unfav_rhp", "Défavorable"),
                       ("scen_moderate_rhp", "Intermédiaire"), ("scen_favor_rhp", "Favorable")]:
        v = parse_pct(cell(key))
        if v is not None:
            scenarios.append(PerformanceScenario(name=label, return_rhp_pct=v))

    sri_raw = parse_number(cell("sri"))
    return KIDData(
        isin=isin,
        product_name=(str(cell("product_name")).strip() if cell("product_name") else None),
        manufacturer=(str(cell("manufacturer")).strip() if cell("manufacturer") else None),
        currency=(str(cell("currency")).strip() if cell("currency") else None),
        sri=int(sri_raw) if sri_raw is not None else None,
        rhp_years=parse_number(cell("rhp_years")),
        costs=costs,
        scenarios=scenarios,
        source_type="ept",
        source_ref=source_ref,
    )


def _iter_rows(path: Path):
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        if not _HAS_XLSX:
            raise RuntimeError("openpyxl requis pour lire un .xlsx (pip install openpyxl)")
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        for r in ws.iter_rows(values_only=True):
            yield [("" if c is None else c) for c in r]
    else:
        # CSV : auto-détection du séparateur (les EPT français sont souvent en ';')
        with open(path, encoding="utf-8-sig", newline="") as f:
            sample = f.read(4096)
            f.seek(0)
            delim = ";" if sample.count(";") >= sample.count(",") else ","
            for r in csv.reader(f, delimiter=delim):
                yield r


def _find_header(rows: list[list]) -> int:
    """L'en-tête EPT n'est pas toujours en ligne 1 (préambule de métadonnées).

    On repère la ligne qui contient à la fois une colonne ISIN et une colonne coût/risque.
    """
    for i, r in enumerate(rows[:30]):
        joined = _norm(" ".join(str(c) for c in r))
        if "isin" in joined and re.search(r"(sri|risk|cost|coût|holding)", joined):
            return i
    return 0


def parse_ept(path: str | Path) -> list[KIDData]:
    """Parse un fichier EPT complet -> une liste de KIDData (un par ISIN/part)."""
    path = Path(path)
    rows = list(_iter_rows(path))
    if not rows:
        return []
    h = _find_header(rows)
    header = [str(c) for c in rows[h]]
    cidx = _build_column_index(header)
    if "isin" not in cidx:
        raise ValueError(
            f"Colonne ISIN introuvable dans {path.name}. "
            f"En-têtes détectés: {header[:12]}… — ajuste COLUMN_MATCHERS/FIELD_CODE_HINTS."
        )
    out = []
    for r in rows[h + 1:]:
        kid = _row_to_kid(r, cidx, str(path))
        if kid:
            out.append(kid)
    return out
